import openpyxl

from flask import flash, redirect, session, request, jsonify

from app import app, db
from .. import utils
from ..database.tables import Product


@app.route('/search_products', methods=['POST'])
def search_products():
    products = Product.query.filter(
        Product.name.ilike(request.form['name'])).all()

    products = [{key: value for key, value in product.__dict__.items() if type(
        value) in [str, float, int]} for product in products]

    if 'id' not in session:
        return jsonify({'products': products, 'id': None}), 200
    return jsonify({'products': products, 'id': session['id']}), 200


@app.route('/get_products/', methods=['GET'])
def get_products():
    products = utils.get_products()
    products = [{key: value for key, value in product.__dict__.items() if type(
        value) in [str, float, int]} for product in products]

    return jsonify(products)


@app.route('/get_product', methods=['POST'])
def get_product():
    product = Product.query.filter_by(id=request.form['id']).first()
    product = {key: value for key, value in product.__dict__.items() if type(value) in [
        str, float, int]}

    return jsonify(product)


@app.route('/admin/add_product', methods=['POST'])
@utils.Admin.authorized_only
def add_product():
    product = Product(**request.form.to_dict())
    db.session.add(product)
    db.session.commit()

    return jsonify(product.name)


@app.route('/admin/add_products', methods=['POST'])
@utils.Admin.authorized_only
def add_products():
    if request.form.get("override", None):
        db.session.query(Product).delete()
    file = request.files['file']
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    row = sheet.max_row + 1
    col = sheet.max_column + 1
    count = 0
    for i in range(2, row):
        curr_item = {}
        for j in range(1, col):
            curr_cell = sheet.cell(row=i, column=j)
            cell_val = curr_cell.value
            curr_item[sheet.cell(row=1, column=j).value] = cell_val
        product = Product.query.filter_by(name=curr_item['name']).first()
        if not product:
            db.session.add(Product(**curr_item))
            count += 1
    db.session.commit()
    flash(f'{count} Products Added')
    return redirect("/admin/manage_products")


@app.route('/admin/delete_product', methods=['DELETE'])
@utils.Admin.authorized_only
def delete_product():
    product = Product.query.filter_by(id=request.form['id']).first()
    if product:
        db.session.delete(product)
        db.session.commit()
        return product.name

    return 'Operation Failed', 400


@app.route('/admin/edit_product', methods=['PUT'])
@utils.Admin.authorized_only
def edit_product():
    product = Product.query.filter_by(id=request.form['id']).first()
    if product:
        product.category = request.form['category']
        product.image = request.form['image']
        product.price = request.form['price']
        product.stock = request.form['stock']
        db.session.commit()
        return 'Product Details Updated', 200
    else:
        return 'Operation Failed', 400
